import openpyxl
import os
def gerarNovaAbaLog(nomeLogs):
    # Diretorio da planilha
    caminhoPath = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../Backup/QP.xlsx'))
    caminho = caminhoPath.replace("\\", "/")

    # Diretorio do log Atualizado

    excel = openpyxl.load_workbook(caminho)
    
    for nomeLog in nomeLogs:
        aba = excel.create_sheet(nomeLog[0])
        logPath = os.path.abspath(os.path.join(os.path.dirname(__file__), f'../../Resultado/LogsAtualizados/{nomeLog[1]}'))
        log = logPath.replace("\\", "/")
        with open(log, 'r') as arquivo:
            linhas = arquivo.readlines()
            for indice, linha in enumerate(linhas, start=1):
                dados = linha.strip().split(';')
                for coluna, valor in enumerate(dados, start=1):
                    aba.cell(row=indice, column=coluna, value=valor)
    
    salvar = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../Resultado/QP.xlsx"))
    excel.save(salvar.replace("\\", "/"))


gerarNovaAbaLog([["SalaExameTecnica", 'Log_salas.txt'],["TuboFluxo",'Log_tubo.txt']])